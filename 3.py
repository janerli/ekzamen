"""дана последовательность n чисел.
Сгенерированных случайным образом.
В .xlsx файл записать только четные числа"""
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import random
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class NumberGenerator:
    def __init__(self, master):
        self.master = master
        master.title("Генератор и запись четных чисел")

        # Фрейм для параметров генерации
        frame_params = ttk.LabelFrame(master, text="Параметры генерации")
        frame_params.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")

        ttk.Label(frame_params, text="Количество чисел:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.count_entry = ttk.Entry(frame_params)
        self.count_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        ttk.Label(frame_params, text="Максимальное число:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.max_entry = ttk.Entry(frame_params)
        self.max_entry.grid(row=1, column=1, padx=5, pady=5, sticky="ew")

        # Кнопка выбора файла
        self.file_label = ttk.Label(master, text="Файл не выбран")
        self.file_label.grid(row=1, column=0, columnspan=2, padx=5, pady=5)
        self.file_button = ttk.Button(master, text="Выбрать файл", command=self.select_file)
        self.file_button.grid(row=2, column=0, columnspan=2, padx=5, pady=5)

        # Кнопка генерации и записи
        self.generate_button = ttk.Button(master, text="Сгенерировать и записать", command=self.generate_and_write, state="disabled")
        self.generate_button.grid(row=3, column=0, columnspan=2, padx=5, pady=5)


        self.file_path = None

    def select_file(self):
       """Открывает диалог сохранения файла и сохраняет путь."""
       file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                               filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")))
       if file_path:
           self.file_path = file_path
           self.file_label.config(text=file_path)
           self.generate_button.config(state="normal")

    def generate_and_write(self):
       """Генерирует числа, фильтрует четные и записывает их в файл .xlsx."""
       try:
           count = int(self.count_entry.get())
           max_value = int(self.max_entry.get())

           if count <= 0 or max_value <= 0:
                messagebox.showerror("Ошибка", "Количество и максимальное значение должны быть положительными.")
                return
           random_numbers = [random.randint(1, max_value) for _ in range(count)]
           even_numbers = [num for num in random_numbers if num % 2 == 0]
           self.write_to_xlsx(even_numbers)
           messagebox.showinfo("Успех", f"Чётные числа записаны в файл: {self.file_path}")

       except ValueError:
           messagebox.showerror("Ошибка", "Введите целые числа в поля количества и максимального значения.")
       except Exception as e:
           messagebox.showerror("Ошибка", f"Произошла ошибка: {e}")

    def write_to_xlsx(self, numbers):
         """Записывает список чисел в файл .xlsx."""
         try:
           workbook = openpyxl.Workbook()
           sheet = workbook.active
           sheet.title = "Четные числа"
           header = ['Четные числа']
           sheet.append(header)

           font = Font(bold=True)
           alignment = Alignment(horizontal='center', vertical='center')
           border = Border(left=Side(border_style='thin'),
                           right=Side(border_style='thin'),
                           top=Side(border_style='thin'),
                           bottom=Side(border_style='thin'))
           for column_num, column_title in enumerate(header):
                column_letter = get_column_letter(column_num+1)
                cell = sheet[f'{column_letter}1']
                cell.font = font
                cell.alignment = alignment
                cell.border = border

           for row_num, num in enumerate(numbers, start=2): # start=2 пропускает строку заголовка
                sheet.append([num])
                cell = sheet[f'A{row_num}']
                cell.border = border
                cell.alignment = alignment
           workbook.save(self.file_path)

         except Exception as e:
             messagebox.showerror("Ошибка", f"Не удалось записать в файл {self.file_path}: {e}")


root = tk.Tk()
generator = NumberGenerator(root)
root.mainloop()
