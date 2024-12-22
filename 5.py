"""найти значение функции в диапазоне (a, b).
Построить график функции и таблицу значений занести в файл .xlsx"""
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import io

def evaluate_function(x):
    """Функция, значения которой мы будем вычислять. (Пример: sin(x))"""
    return np.sin(x)

def create_plot(min_x, max_x):
    """Создает график функции в заданном диапазоне и возвращает его как объект plt.Figure."""
    x = np.linspace(min_x, max_x, 400)
    y = [evaluate_function(val) for val in x]

    fig, ax = plt.subplots(figsize=(8, 6))
    ax.plot(x, y)
    ax.set_title("График функции f(x) = sin(x)")
    ax.set_xlabel("x")
    ax.set_ylabel("f(x)")
    ax.grid(True)
    return fig

def save_to_excel(min_x, max_x):
    """Сохраняет таблицу значений функции и график в Excel файл."""
    x_values = np.linspace(min_x, max_x, 400)
    y_values = [evaluate_function(val) for val in x_values]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Значения функции"

    # Заголовки
    sheet['A1'] = 'x'
    sheet['B1'] = 'f(x)'

    # Запись значений
    for row, (x, y) in enumerate(zip(x_values, y_values), start=2):
        sheet.cell(row=row, column=1, value=x)
        sheet.cell(row=row, column=2, value=y)

    # Выравнивание по ширине столбца
    for column in range(1, 3):
        max_length = 0
        column_letter = get_column_letter(column)
        for cell in sheet[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Добавление графика
    fig = create_plot(min_x, max_x)
    img_buf = io.BytesIO()
    fig.savefig(img_buf, format='png')
    img = Image(img_buf)
    sheet.add_image(img, 'D1')

    # Сохранение файла
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")]
    )
    if file_path:
        try:
            workbook.save(file_path)
            messagebox.showinfo("Успех", "Данные и график успешно сохранены в Excel файл.")
        except Exception as e:
             messagebox.showerror("Ошибка", f"Не удалось сохранить файл: {e}")
    plt.close(fig) # Закрываем фигуру matplotlib


def process_function():
    """Получает диапазон и сохраняет в Excel."""
    try:
        min_x = float(min_entry.get())
        max_x = float(max_entry.get())

        if min_x >= max_x:
            messagebox.showerror("Ошибка", "Минимальное значение должно быть меньше максимального")
            return

        save_to_excel(min_x, max_x)
    except ValueError:
        messagebox.showerror("Ошибка", "Пожалуйста, введите числа в поля")

# Создание главного окна
window = tk.Tk()
window.title("Вычисление функции")
window.geometry("500x300")

# Фрейм для ввода диапазона
range_frame = ttk.Frame(window)
range_frame.pack(pady=10)

ttk.Label(range_frame, text="Минимальное значение (a):").grid(row=0, column=0, padx=5, pady=5)
min_entry = ttk.Entry(range_frame)
min_entry.grid(row=0, column=1, padx=5, pady=5)
min_entry.insert(0, "0") # Пример значения

ttk.Label(range_frame, text="Максимальное значение (b):").grid(row=1, column=0, padx=5, pady=5)
max_entry = ttk.Entry(range_frame)
max_entry.grid(row=1, column=1, padx=5, pady=5)
max_entry.insert(0, "10") # Пример значения

# Кнопка обработки
process_button = ttk.Button(window, text="Вычислить и сохранить в Excel", command=process_function)
process_button.pack(pady=10)


window.mainloop()
